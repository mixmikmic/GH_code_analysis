import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
get_ipython().run_line_magic('matplotlib', 'inline')

from bokeh.plotting import figure, output_file, show, ColumnDataSource
from bokeh.io import output_notebook
from bokeh.models import HoverTool, Label, LabelSet, Span

TOOLS = " crosshair, pan, wheel_zoom, zoom_in, zoom_out, box_zoom, undo, redo, reset, tap, save, box_select, poly_select, lasso_select"
output_notebook()

df = pd.read_excel(r"C:\EY\Thought Leadership\MA Correlation\correlation_test_data.xlsx" , sheetname='Sheet1')

file_path = r"C:\EY\Thought Leadership\MA Correlation\correlation_matrix.xlsx"
try:
    book = load_workbook(file_path)
except:
    wb = Workbook()
    wb.save(file_path)
    book = load_workbook(file_path)
writer = pd.ExcelWriter(file_path, engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

# df.columns.tolist()

dataset = df[['Company',
 'Company Ticker',
 'market_cap',
 'Sector',
 'Est.Periods',
 'avg_TSR',
 'avg_ROA',
 'avg_ROE',
 'avg_EPS',
 'avg_dividend_ratio',
 'Rev_grow.Avg',
 'ebit_grow.avg',
 'avg_cost_growh',
 'Acquisitions',
 'ma_per_year',
 'Acq value',
 'Divestment',
 'div_per_year',
 'div_value',
 'Div as % deal value',
 'Deal intensity',
 'avg_deal_size',
 'avg_deal_size_mktcap',
 'domestic',
 'cross_border',
 'domestic_percentage',
 'in_sector',
 'out_sector',
 'sector_percentage'
 ]]

dataset.head()

dataset.dtypes

dataset.describe()

dataset = dataset.fillna(value=0)

dataset.shape

dataset = dataset[~(((dataset['div_value']==0) & (dataset['Divestment']!=0)) | ((dataset['div_value']!=0) & (dataset['Divestment']==0)))] 

dataset = dataset[~(((dataset['Acq value']==0) & (dataset['Acquisitions']!=0)) | ((dataset['Acq value']!=0) & (dataset['Acquisitions']==0)))]

dataset = dataset[~(((dataset['div_per_year']==0) & (dataset['Divestment']!=0)) | ((dataset['div_per_year']!=0) & (dataset['Divestment']==0)))] 

dataset = dataset[~(((dataset['ma_per_year']==0) & (dataset['Acquisitions']!=0)) | ((dataset['ma_per_year']!=0) & (dataset['Acquisitions']==0)))] 

dataset = dataset[np.abs(dataset['avg_TSR']-df['avg_TSR'].mean())<=(2*dataset['avg_TSR'].std())]

dataset = dataset[np.abs(dataset['avg_ROA']-df['avg_ROA'].mean())<=(2*dataset['avg_ROA'].std())]

dataset = dataset[np.abs(dataset['avg_ROE']-df['avg_ROE'].mean())<=(2*dataset['avg_ROE'].std())]

dataset = dataset[np.abs(dataset['avg_cost_growh']-df['avg_cost_growh'].mean())<=(2*dataset['avg_cost_growh'].std())]

dataset = dataset[np.abs(dataset['avg_deal_size_mktcap']-df['avg_deal_size_mktcap'].mean())<=(2*dataset['avg_deal_size_mktcap'].std())]

dataset = dataset[np.abs(dataset['Deal intensity']-df['Deal intensity'].mean())<=(2*dataset['Deal intensity'].std())]

dataset = dataset[np.abs(dataset['Rev_grow.Avg']-df['Rev_grow.Avg'].mean())<=(2*dataset['Rev_grow.Avg'].std())]

dataset = dataset[np.abs(dataset['ebit_grow.avg']-df['ebit_grow.avg'].mean())<=(2*dataset['ebit_grow.avg'].std())]

dataset.to_excel(writer, 'Clean Data', index=False)
writer.save()

dataset.shape

dataset.describe()

organic = dataset[(dataset['Acquisitions']==0)]
in_organic = dataset[((dataset['Acquisitions']!=0))]
organic.shape, in_organic.shape

organic.describe()

organic.corr()

in_organic.describe()

in_organic.corr()

# Output the descriptive statistics and correlation matrix
organic.describe().to_excel(writer, 'organic_describe')
organic_corr = organic.corr().dropna(how='any', thresh=5, axis=0)
organic_corr = organic_corr.dropna(how='any', thresh=5, axis=1)
organic_corr.to_excel(writer, 'organic_corr')
writer.save()

plt.figure(figsize=(12,8))
sns.heatmap(organic_corr)

# Output the scatter plot of TSR and de

# Output the descriptive statistics and correlation matrix
in_organic.describe().to_excel(writer, 'in_organic_describe')
in_organic_corr = in_organic.corr().dropna(how='any', thresh=5, axis=0)
in_organic_corr = in_organic_corr.dropna(how='any', thresh=5, axis=1)
in_organic_corr.to_excel(writer, 'in_organic_corr')
writer.save()

plt.figure(figsize=(12,8))
sns.heatmap(in_organic_corr)

# Output the scatter plot of TSR and deal intensity of inorganic growths.
in_organic.loc[:,'bubble_size'] = in_organic['ma_per_year']*10
source1 = ColumnDataSource(in_organic)
source2 = ColumnDataSource(organic)
in_organic_plot = figure(tools=TOOLS, width=900, height=500, x_axis_label="Deal Intensity", y_axis_label="TSR", x_range=(0,15))
in_organic_plot.scatter(y='avg_TSR', x='Deal intensity', size=10, alpha=0.6, source=source1)
#in_organic_plot.scatter(y='avg_TSR', x='Deal intensity', alpha=0.6, color='blue', source=source2)
hover = HoverTool(tooltips=[("Company", "@Company"), ("ma_per_year", "@ma_per_year"), ("Average TSR", "@avg_TSR")])
in_organic_plot.add_tools(hover)
in_organic_plot.ray(x=0, y=organic['avg_TSR'].mean(), angle=0, color='red', legend="Organic Avg. TSR", length=900, )
in_organic_plot.ray(x=0, y=in_organic['avg_TSR'].mean(), angle=0, color='green', legend="In Organic Avg. TSR", length=900)
organic_label = Label(x=800, y=-0.05, x_units='screen', text=str('{0:.2f}'.format(organic['avg_TSR'].mean()*100)) + "%", text_font_size='10pt')
in_organic_label = Label(x=800, y=0.05, x_units='screen', text=str('{0:.2f}'.format(in_organic['avg_TSR'].mean()*100)) + "%", text_font_size='10pt')
in_organic_plot.add_layout(organic_label)
in_organic_plot.add_layout(in_organic_label)
show(in_organic_plot)

dataset.shape

non_acquirer = dataset[(dataset['Acquisitions']==0)]
periodic_acquirer = dataset[(dataset['ma_per_year']<=0.2) & (dataset['ma_per_year']>0)]
frequent_acquirer = dataset[(dataset['ma_per_year'])>0.2]
non_acquirer.shape, periodic_acquirer.shape, frequent_acquirer.shape

periodic_acquirer.describe()

periodic_acquirer.corr()

frequent_acquirer.describe()

frequent_acquirer.corr()

periodic_acquirer.describe().to_excel(writer, 'period_describe')
periodic_acquirer_corr = periodic_acquirer.corr().dropna(how='any', thresh=5, axis=0)
periodic_acquirer_corr = periodic_acquirer_corr.dropna(how='any', thresh=5, axis=0)
periodic_acquirer_corr.to_excel(writer, 'periodic_corr')
writer.save()

frequent_acquirer.describe().to_excel(writer, 'frequent_describe')
frequent_corr = frequent_acquirer.corr().dropna(how='any', thresh=5, axis=0)
frequent_corr = frequent_corr.dropna(how='any', thresh=5, axis=0)
frequent_corr.to_excel(writer, 'frequent_corr')
writer.save()

source = ColumnDataSource(dataset)
source1 = ColumnDataSource(organic)
source2 = ColumnDataSource(periodic_acquirer)
source3 = ColumnDataSource(frequent_acquirer)
dataset_plot = figure(tools=TOOLS, width=900, height=500, x_axis_label="Deal Frequency", y_axis_label="TSR", x_range=(0,10))
dataset_plot.scatter(y='avg_TSR', x='ma_per_year', size=10, alpha=0.6, color='green', source=source1)
dataset_plot.scatter(y='avg_TSR', x='ma_per_year', size=10, alpha=0.6, color='blue', source=source2)
dataset_plot.scatter(y='avg_TSR', x='ma_per_year', size=10, alpha=0.6, color='gray', source=source3)
hover = HoverTool(tooltips=[("Company", "@Company"), ("ma_per_year", "@ma_per_year"), ("Average TSR", "@avg_TSR")])
dataset_plot.ray(x=0, y=organic['avg_TSR'].mean(), angle=0, color='red', legend="Organic Avg. TSR", length=900)
dataset_plot.ray(x=0, y=periodic_acquirer['avg_TSR'].mean(), angle=0, color='green', legend="Periodic Acquirer's Avg. TSR", length=900)
dataset_plot.ray(x=0, y=frequent_acquirer['avg_TSR'].mean(), angle=0, color='gray', legend="Frequent Acquirer's Avg. TSR", length=900)
organic_label = Label(x=800, y=-0.05, x_units='screen', text=str('{0:.2f}'.format(organic['avg_TSR'].mean()*100)) + "%", text_font_size='10pt')
periodic_label = Label(x=800, y=0.01, x_units='screen', text=str('{0:.2f}'.format(periodic_acquirer['avg_TSR'].mean()*100)) + "%", text_font_size='10pt')
frequent_label = Label(x=800, y=0.05, x_units='screen', text=str('{0:.2f}'.format(frequent_acquirer['avg_TSR'].mean()*100)) + "%", text_font_size='10pt')
dataset_plot.add_tools(hover)
dataset_plot.add_layout(organic_label)
dataset_plot.add_layout(periodic_label)
dataset_plot.add_layout(frequent_label)
show(dataset_plot)

dividend_stock = dataset[dataset['avg_dividend_ratio']>0.1]
growth_stock = dataset[~(dataset['avg_dividend_ratio']>0.1)]
dividend_stock.shape, growth_stock.shape

dividend_stock.describe()

dividend_stock.corr()

growth_stock.describe()

growth_stock.corr()

dividend_stock.describe().to_excel(writer, 'dividend_describe')
dividend_corr = dividend_stock.corr().dropna(how='any', thresh=5, axis=0)
dividend_corr = dividend_corr.dropna(how='any', thresh=5, axis=0)
dividend_corr.to_excel(writer, 'dividend_corr')
writer.save()

growth_stock.describe().to_excel(writer, 'growth_describe')
growth_corr = growth_stock.corr().dropna(how='any', thresh=5, axis=0)
growth_corr = growth_corr.dropna(how='any', thresh=5, axis=0)
growth_corr.to_excel(writer, 'growth_corr')
writer.save()

dividend_stock[dividend_stock['Company'].str.contains("Steadfast")]

source_1 = ColumnDataSource(dividend_stock)
source_2 = ColumnDataSource(growth_stock)
dataset_plot = figure(tools=TOOLS, width=900, height=500, x_axis_label="Deal Frequency", y_axis_label="TSR", x_range=(0,15))
dataset_plot.scatter(y='avg_TSR', x='ma_per_year', size=10, alpha=0.5, color='red', source=source_1, legend='Dividend Stocks')
dataset_plot.scatter(y='avg_TSR', x='ma_per_year', size=10, alpha=0.5, color='blue', source=source_2, legend='Growth Stocks')
dataset_plot.ray(x=0, y=dividend_stock['avg_TSR'].mean(), angle=0, color='green', legend="Dividend Stock Avg. TSR", length=900)
dataset_plot.ray(x=0, y=growth_stock['avg_TSR'].mean(), angle=0, color='red', legend="Growth Stock Avg. TSR", length=900)
hover = HoverTool(tooltips=[("Company", "@Company"), ("ma_per_year", "@ma_per_year"), ("Average TSR", "@avg_TSR")])
dataset_plot.add_tools(hover)
dividend_label = Label(x=800, y=0.07, x_units='screen', text=str('{0:.2f}'.format(dividend_stock['avg_TSR'].mean()*100)) + "%", text_font_size='10pt')
growth_label = Label(x=800, y=-0.01, x_units='screen', text=str('{0:.2f}'.format(growth_stock['avg_TSR'].mean()*100)) + "%", text_font_size='10pt')
dataset_plot.add_layout(dividend_label)
dataset_plot.add_layout(growth_label)
show(dataset_plot)

domestic_acquirer = dataset[(dataset['domestic_percentage']>0.8)]
cross_border_acquirer = dataset[(dataset['domestic_percentage']<0.8) & (dataset['Acquisitions']!=0)]

domestic_acquirer.shape, cross_border_acquirer.shape

domestic_acquirer.describe()

domestic_acquirer.corr()

cross_border_acquirer.describe()

cross_border_acquirer.corr()

domestic_acquirer.describe().to_excel(writer, 'domestic_describe')
domestic_corr = domestic_acquirer.corr().dropna(how='any', thresh=5, axis=0)
domestic_corr = domestic_corr.dropna(how='any', thresh=5, axis=0)
domestic_corr.to_excel(writer, 'domestic_corr')
writer.save()

cross_border_acquirer.describe().to_excel(writer, 'cross_border_describe')
cross_border_corr = cross_border_acquirer.corr().dropna(how='any', thresh=5, axis=0)
cross_border_corr = cross_border_corr.dropna(how='any', thresh=5, axis=0)
cross_border_corr.to_excel(writer, 'cross_border_corr')
writer.save()

source_1 = ColumnDataSource(domestic_acquirer)
source_2 = ColumnDataSource(cross_border_acquirer)
dataset_plot = figure(tools=TOOLS, width=900, height=500, x_axis_label="% of Domestic Acquisitions (in $ terms)", y_axis_label="TSR", x_range=(0,8))
dataset_plot.scatter(y='avg_TSR', x='ma_per_year', size=10, alpha=0.6, color='red', source=source_1, legend='Predominantly Domestic Acquirers')
dataset_plot.scatter(y='avg_TSR', x='ma_per_year', size=10, alpha=0.6, color='blue', source=source_2, legend='Predominantly Cross_border Acquirers')
dataset_plot.ray(x=0, y=domestic_acquirer['avg_TSR'].mean(), angle=0, color='green', legend="Predominantly Domestic Acquirers Avg. TSR", length=900)
dataset_plot.ray(x=0, y=cross_border_acquirer['avg_TSR'].mean(), angle=0, color='gray', legend="Predominantly Cross Border Avg. TSR", length=900)
hover = HoverTool(tooltips=[("Company", "@Company"), ("ma_per_year", "@ma_per_year"), ("Average TSR", "@avg_TSR")])
dataset_plot.add_tools(hover)
domestic_label = Label(x=600, y=0.04, x_units='screen', text=str('{0:.2f}'.format(domestic_acquirer['avg_TSR'].mean()*100)) + "%", text_font_size='10pt')
cross_border_label = Label(x=600, y=-0.03, x_units='screen', text=str('{0:.2f}'.format(cross_border_acquirer['avg_TSR'].mean()*100)) + "%", text_font_size='10pt')
dataset_plot.add_layout(domestic_label)
dataset_plot.add_layout(cross_border_label)
show(dataset_plot)

sector_acquirer = dataset[(dataset['sector_percentage']>0.8)]
diversified_acquirer = dataset[(dataset['sector_percentage']<0.8) & (dataset['Acquisitions']!=0)]
sector_acquirer.shape, diversified_acquirer.shape

sector_acquirer.describe()

sector_acquirer.corr()

diversified_acquirer.describe()

diversified_acquirer.corr()

sector_acquirer.describe().to_excel(writer, 'in_sector_describe')
sector_corr = sector_acquirer.corr().dropna(how='any', thresh=5, axis=0)
sector_corr = sector_corr.dropna(how='any', thresh=5, axis=0)
sector_corr.to_excel(writer, 'in_sector_corr')
writer.save()

diversified_acquirer.describe().to_excel(writer, 'diversified_describe')
diversified_corr = diversified_acquirer.corr().dropna(how='any', thresh=5, axis=0)
diversified_corr = diversified_corr.dropna(how='any', thresh=5, axis=0)
diversified_corr.to_excel(writer, 'diversified_corr')
writer.save()

source_1 = ColumnDataSource(sector_acquirer)
source_2 = ColumnDataSource(diversified_acquirer)
dataset_plot = figure(tools=TOOLS, width=900, height=500, x_axis_label="Deal Frequency (Transactions per Year)", y_axis_label="TSR", x_range=(0,8))
dataset_plot.scatter(y='avg_TSR', x='ma_per_year', size=10, alpha=0.6, color='red', source=source_1, legend='Predominantly Industrial Acquirers')
dataset_plot.scatter(y='avg_TSR', x='ma_per_year', size=10, alpha=0.6, color='blue', source=source_2, legend='Predominantly Diversified Acquirers')
dataset_plot.ray(x=0, y=sector_acquirer['avg_TSR'].mean(), angle=0, color='green', legend="Predominantly Industrial Acquirers Avg. TSR", length=900)
dataset_plot.ray(x=0, y=diversified_acquirer['avg_TSR'].mean(), angle=0, color='gray', legend="Predominantly Diversified Avg. TSR", length=900)
hover = HoverTool(tooltips=[("Company", "@Company"), ("ma_per_year", "@ma_per_year"), ("Average TSR", "@avg_TSR")])
dataset_plot.add_tools(hover)
sector_label = Label(x=600, y=0.04, x_units='screen', text=str('{0:.2f}'.format(sector_acquirer['avg_TSR'].mean()*100)) + "%", text_font_size='10pt')
diversified_label = Label(x=600, y=-0.03, x_units='screen', text=str('{0:.2f}'.format(diversified_acquirer['avg_TSR'].mean()*100)) + "%", text_font_size='10pt')
dataset_plot.add_layout(sector_label)
dataset_plot.add_layout(diversified_label)
show(dataset_plot)

dataset.describe()

good_stocks = dataset[dataset['avg_TSR']>0.087]
bad_stocks = dataset[(dataset['avg_TSR']>-0.17) & (dataset['avg_TSR']<0)]
ugly_stocks = dataset[dataset['avg_TSR']<-0.17]

good_stocks.shape, bad_stocks.shape, ugly_stocks.shape

source_1 = ColumnDataSource(good_stocks)
source_2 = ColumnDataSource(bad_stocks)
source_3 = ColumnDataSource(ugly_stocks)
dataset_plot = figure(tools=TOOLS, width=900, height=500, x_axis_label="Deal Frequency", y_axis_label="TSR", x_range=(0,8))
dataset_plot.scatter(y='avg_TSR', x='ma_per_year', size=10, alpha=0.6, color='green', source=source_1, legend='The Good')
dataset_plot.scatter(y='avg_TSR', x='ma_per_year', size=10, alpha=0.6, color='orange', source=source_2, legend='The Bad')
dataset_plot.scatter(y='avg_TSR', x='ma_per_year', size=10, alpha=0.6, color='red', source=source_3, legend='The Ugly')
dataset_plot.ray(x=0, y=good_stocks['avg_TSR'].mean(), angle=0, color='green', legend="The Good's Avg. TSR", length=900)
dataset_plot.ray(x=0, y=bad_stocks['avg_TSR'].mean(), angle=0, color='orange', legend="The Bad's Avg. TSR", length=900)
dataset_plot.ray(x=0, y=ugly_stocks['avg_TSR'].mean(), angle=0, color='red', legend="The Ugly's Avg. TSR", length=900)
hover = HoverTool(tooltips=[("Company", "@Company"), ("ma_per_year", "@ma_per_year"), ("Average TSR", "@avg_TSR")])
dataset_plot.add_tools(hover)
show(dataset_plot)

good_acquirer = good_stocks[(good_stocks['Acquisitions']!=0) | (good_stocks['Divestment']!=0)]
bad_acquirer = bad_stocks[(bad_stocks['Acquisitions']!=0) | (bad_stocks['Divestment']!=0)]
ugly_acquirer = ugly_stocks[(ugly_stocks['Acquisitions']!=0) | (ugly_stocks['Divestment']!=0)]

good_acquirer.shape, bad_acquirer.shape, ugly_acquirer.shape

good_acquirer.loc[:, 'bubble_size'] = good_acquirer['ma_per_year']*10
bad_acquirer.loc[:, 'bubble_size'] = bad_acquirer['ma_per_year']*10
ugly_acquirer.loc[:, 'bubble_size'] = ugly_acquirer['ma_per_year']*10

source_1 = ColumnDataSource(good_acquirer)
source_2 = ColumnDataSource(bad_acquirer)
source_3 = ColumnDataSource(ugly_acquirer)
dataset_plot = figure(tools=TOOLS, width=900, height=500, x_axis_label="Deal Intensity", y_axis_label="TSR", x_range=(0,8))
dataset_plot.scatter(y='avg_TSR', x='Deal intensity', size='bubble_size', alpha=0.6, color='green', source=source_1, legend='The Good')
dataset_plot.scatter(y='avg_TSR', x='Deal intensity', size='bubble_size', alpha=0.6, color='orange', source=source_2, legend='The Bad')
dataset_plot.scatter(y='avg_TSR', x='Deal intensity', size='bubble_size', alpha=0.6, color='red', source=source_3, legend='The Ugly')
dataset_plot.ray(x=0, y=good_acquirer['avg_TSR'].mean(), angle=0, color='green', legend="The Good's Avg. TSR", length=900)
dataset_plot.ray(x=0, y=bad_acquirer['avg_TSR'].mean(), angle=0, color='orange', legend="The Bad's Avg. TSR", length=900)
dataset_plot.ray(x=0, y=ugly_acquirer['avg_TSR'].mean(), angle=0, color='red', legend="The Ugly's Avg. TSR", length=900)
hover = HoverTool(tooltips=[("Company", "@Company"), ("ma_per_year", "@ma_per_year"), ("Average TSR", "@avg_TSR")])
dataset_plot.add_tools(hover)
good_label = Label(x=600, y=0.28, x_units='screen', text=str('{0:.2f}'.format(good_acquirer['avg_TSR'].mean()*100)) + "%", text_font_size='10pt')
bad_label = Label(x=600, y=-0.07, x_units='screen', text=str('{0:.2f}'.format(bad_acquirer['avg_TSR'].mean()*100)) + "%", text_font_size='10pt')
ugly_label = Label(x=600, y=-0.29, x_units='screen', text=str('{0:.2f}'.format(ugly_acquirer['avg_TSR'].mean()*100)) + "%", text_font_size='10pt')
dataset_plot.add_layout(good_label)
dataset_plot.add_layout(bad_label)
dataset_plot.add_layout(ugly_label)
show(dataset_plot)

good_acquirer['Company'].tolist()

good_acquirer.columns

good_acquirer['rev_vs_cost'] = good_acquirer['Rev_grow.Avg']/good_acquirer['avg_cost_growh']
bad_acquirer['rev_vs_cost'] = bad_acquirer['Rev_grow.Avg']/bad_acquirer['avg_cost_growh']
ugly_acquirer['rev_vs_cost'] = ugly_acquirer['Rev_grow.Avg']/ugly_acquirer['avg_cost_growh']
good_acquirer.replace([np.inf, -np.inf], np.nan, inplace=True)
bad_acquirer.replace([np.inf, -np.inf], np.nan, inplace=True)
ugly_acquirer.replace([np.inf, -np.inf], np.nan, inplace=True)

good_acquirer.columns

good_acquirer['ebit_vs_cost'] = good_acquirer['ebit_grow.avg']/good_acquirer['avg_cost_growh']
bad_acquirer['ebit_vs_cost'] = bad_acquirer['ebit_grow.avg']/bad_acquirer['avg_cost_growh']
ugly_acquirer['ebit_vs_cost'] = ugly_acquirer['ebit_grow.avg']/ugly_acquirer['avg_cost_growh']
organic['ebit_vs_cost'] = organic['ebit_grow.avg']/organic['avg_cost_growh']
good_acquirer.replace([np.inf, -np.inf], np.nan, inplace=True)
bad_acquirer.replace([np.inf, -np.inf], np.nan, inplace=True)
ugly_acquirer.replace([np.inf, -np.inf], np.nan, inplace=True)
organic.replace([np.inf, -np.inf], np.nan, inplace=True)

good_acquirer['rev_vs_cost'].mean(), bad_acquirer['rev_vs_cost'].mean(), ugly_acquirer['rev_vs_cost'].mean()

good_acquirer['ebit_vs_cost'].mean(), bad_acquirer['ebit_vs_cost'].mean(), ugly_acquirer['ebit_vs_cost'].mean(), organic['ebit_vs_cost'].mean()

source_1 = ColumnDataSource(good_acquirer)
source_2 = ColumnDataSource(bad_acquirer)
source_3 = ColumnDataSource(ugly_acquirer)
source_4 = ColumnDataSource(organic)
dataset_plot = figure(tools=TOOLS, width=900, height=500, x_axis_label="EBIT Growth Rate over Cost Growth Rate", y_axis_label="TSR", x_range=(-16, 10), y_range=(-1,1))
dataset_plot.scatter(y='avg_TSR', x='ebit_vs_cost', size='Deal intensity', alpha=0.6, color='green', source=source_1, legend='The Good')
dataset_plot.scatter(y='avg_TSR', x='ebit_vs_cost', size='Deal intensity', alpha=0.6, color='orange', source=source_2, legend='The Bad')
dataset_plot.scatter(y='avg_TSR', x='ebit_vs_cost', size='Deal intensity', alpha=0.6, color='red', source=source_3, legend='The Ugly')
dataset_plot.scatter(y='avg_TSR', x='ebit_vs_cost', size=10, alpha=0.6, color='grey', source=source_4, legend='The Organic')
good_vline = Span(location=good_acquirer['ebit_vs_cost'].mean(), dimension='height', line_color='green')
bad_vline = Span(location=bad_acquirer['ebit_vs_cost'].mean(), dimension='height', line_color='orange')
ugly_vline = Span(location=ugly_acquirer['ebit_vs_cost'].mean(), dimension='height', line_color='red')
organic_vline = Span(location=organic['ebit_vs_cost'].mean(), dimension='height', line_color='gray')
# dataset_plot.ray(x=good_acquirer['ebit_vs_cost'].mean(), y=-1, angle=0, color='green', legend="The Good's Avg. TSR", length=900)
# dataset_plot.ray(x=-1, y=bad_acquirer['ebit_vs_cost'].mean(), angle=0, color='orange', legend="The Bad's Avg. TSR", length=900)
# dataset_plot.ray(x=-1, y=ugly_acquirer['ebit_vs_cost'].mean(), angle=0, color='red', legend="The Ugly's Avg. TSR", length=900)
# dataset_plot.line([good_acquirer['ebit_vs_cost'].mean(), -1],[-1, 1])
hover = HoverTool(tooltips=[("Company", "@Company"), ("ma_per_year", "@ma_per_year"), ("Average TSR", "@avg_TSR")])
dataset_plot.add_tools(hover)
dataset_plot.add_layout(good_vline)
dataset_plot.add_layout(bad_vline)
dataset_plot.add_layout(ugly_vline)
dataset_plot.add_layout(organic_vline)
good_label = Label(x=650, y=0.7, x_units='screen', text=str('{0:.2f}'.format(good_acquirer['ebit_vs_cost'].mean())), text_font_size='10pt')
bad_label = Label(x=50, y=0.7, x_units='screen', text=str('{0:.2f}'.format(bad_acquirer['ebit_vs_cost'].mean())), text_font_size='10pt')
ugly_label = Label(x=500, y=0.7, x_units='screen', text=str('{0:.2f}'.format(ugly_acquirer['ebit_vs_cost'].mean())) , text_font_size='10pt')
dataset_plot.add_layout(good_label)
dataset_plot.add_layout(bad_label)
dataset_plot.add_layout(ugly_label)
show(dataset_plot)



