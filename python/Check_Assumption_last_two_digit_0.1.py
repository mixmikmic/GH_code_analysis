from __future__ import division, generators, absolute_import, print_function
import numpy as np
from scipy.stats import binom
from openpyxl import load_workbook

DATA_DIR = '../data/'
wb = load_workbook(DATA_DIR + 'Bishayee Coulter Counts.10.20.97-7.16.01.xlsx') 

ws = wb.get_sheet_by_name('Sheet1')
rawData = []

for i in range(3, 1732):
    for col in ['C','D','E']:
        try:
            rawData.append(float(ws[col + str(i)].value))
        except:
            # if empty cell, report location in spreadsheet
            print(col,i)

# this calcualte the number of all high-digit (more than 3) numbers
high_digit = [int(i) for i in rawData if i > 99]
print(len(high_digit))
# this calculate the number among all high_digit numbers, that last digit is 1 smaller than 2-nd last digit
# i.e. x10, xx21, xx32, xx43, xx54, xx65, xx76, xx87, xx98, xx09
interested_digit = [i for i in high_digit if int((i % 100) / 10) == ((i % 10 + 1) % 10)]
print(len(interested_digit))

print('proportion of such pairs in RTS Coulter')
print(len(interested_digit)/len(high_digit))
print('p-values of such pairs in RTS Coulter')
print(1 - binom.cdf(k=len(interested_digit)-1, n=len(high_digit), p=0.1, loc=0))



DATA_DIR = '../data/'
wb = load_workbook(DATA_DIR + 'Other Investigators in Lab.Coulter Counts.4.15.92-5.21.05.xlsx') 

ws = wb.get_sheet_by_name('Sheet1')
rawData = []

for i in range(3, 1010):
    for col in ['C','D','E']:
        try:
            rawData.append(float(ws[col + str(i)].value))
        except:
            # if empty cell, report location in spreadsheet
            print(col,i)

# this calcualte the number of all high-digit (more than 3) numbers
high_digit = [int(i) for i in rawData if i > 99]
print(len(high_digit))
# this calculate the number among all high_digit numbers, that last digit is 1 smaller than 2-nd last digit
# i.e. x10, xx21, xx32, xx43, xx54, xx65, xx76, xx87, xx98, xx09
interested_digit = [i for i in high_digit if int((i % 100) / 10) == ((i % 10 + 1) % 10)]
print(len(interested_digit))

print('proportion of such pairs in Others Coulter')
print(len(interested_digit)/len(high_digit))
print('p-values of such pairs in Others Coulter')
print(1 - binom.cdf(k=len(interested_digit)-1, n=len(high_digit), p=0.1, loc=0))



DATA_DIR = '../data/'
wb = load_workbook(DATA_DIR + 'Bishayee Colony Counts 10.27.97-3.8.01.xlsx') 
ws = wb.get_sheet_by_name('Sheet1')
rawData = []

for i in range(4, 1366):
    for col in ['D','E','F']:
        try:
            rawData.append(float(ws[col + str(i)].value))
        except:
            # if empty cell, report location in spreadsheet
            print(col,i)

# this calcualte the number of all high-digit (more than 3) numbers
high_digit = [int(i) for i in rawData if i > 99]
print(len(high_digit))
# this calculate the number among all high_digit numbers, that last digit is 1 smaller than 2-nd last digit
# i.e. x10, xx21, xx32, xx43, xx54, xx65, xx76, xx87, xx98, xx09
interested_digit = [i for i in high_digit if int((i % 100) / 10) == ((i % 10 + 1) % 10)]
print(len(interested_digit))

print('proportion of such pairs in RTS Colony')
print(len(interested_digit)/len(high_digit))
print('p-values of such pairs in RTS Colony')
print(1 - binom.cdf(k=len(interested_digit)-1, n=len(high_digit), p=0.1, loc=0))



DATA_DIR = '../data/'
wb = load_workbook(DATA_DIR + 'Other Investigators in Lab.Colony Counts.4.23.92-11.27.02.xlsx') 

ws = wb.get_sheet_by_name('Sheet1')
rawData = []

for i in range(3, 626):
    for col in ['D','E','F']:
        try:
            rawData.append(float(ws[col + str(i)].value))
        except:
            # if empty cell, report location in spreadsheet
            print(col,i)

# this calcualte the number of all high-digit (more than 3) numbers
high_digit = [int(i) for i in rawData if i > 99]
print(len(high_digit))
# this calculate the number among all high_digit numbers, that last digit is 1 smaller than 2-nd last digit
# i.e. x10, xx21, xx32, xx43, xx54, xx65, xx76, xx87, xx98, xx09
interested_digit = [i for i in high_digit if int((i % 100) / 10) == ((i % 10 + 1) % 10)]
print(len(interested_digit))

print('proportion of such pairs in Others Colony')
print(len(interested_digit)/len(high_digit))
print('p-values of such pairs in Others Colony')
print(1 - binom.cdf(k=len(interested_digit)-1, n=len(high_digit), p=0.1, loc=0))



